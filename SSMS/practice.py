import openpyxl

def divide_columns_by_design_area(file_path):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = {cell.value: cell.column for cell in sheet[1]}

    kn_columns = ['C_25KN', 'C_30KN', 'C_40KN', 'C_50KN', 'C_60KN', 'C_70KN', 'C_80KN']
    design_area1_col = headers['Design Area 1']
    design_area2_col = headers['Design Area 2']
    
    for row in sheet.iter_rows(min_row=2, values_only=False):
        try:
            design_area1 = float(row[design_area1_col - 1].value)
            design_area2 = float(row[design_area2_col - 1].value)
        except (TypeError, ValueError):
            print(f"Skipping row {row[0].row} due to invalid Design Area values")
            continue

        for kn_col in kn_columns:
            kn_col_index = headers.get(kn_col)
            if kn_col_index is not None:
                try:
                    kn_value = float(row[kn_col_index - 1].value)
                    kn_value_divided_by_area1 = kn_value / design_area1
                    kn_value_divided_by_area2 = kn_value / design_area2

                    print(f"Row {row[0].row} - {kn_col} / Design Area 1: {kn_value_divided_by_area1}")
                    print(f"Row {row[0].row} - {kn_col} / Design Area 2: {kn_value_divided_by_area2}")

                    new_col_name_area1 = kn_col.replace('C_', '') + 'KN / DA1'
                    new_col_name_area2 = kn_col.replace('C_', '') + 'KN / DA2'
                    
                    if new_col_name_area1 not in headers:
                        headers[new_col_name_area1] = len(headers) + 1
                        sheet.cell(row=1, column=headers[new_col_name_area1], value=new_col_name_area1)
                    
                    if new_col_name_area2 not in headers:
                        headers[new_col_name_area2] = len(headers) + 1
                        sheet.cell(row=1, column=headers[new_col_name_area2], value=new_col_name_area2)
                    
                    sheet.cell(row=row[0].row, column=headers[new_col_name_area1], value=kn_value_divided_by_area1)
                    sheet.cell(row=row[0].row, column=headers[new_col_name_area2], value=kn_value_divided_by_area2)
                except (TypeError, ValueError):
                    print(f"Skipping row {row[0].row} due to invalid {kn_col} value")
                    continue
    workbook.save(file_path)

if __name__ == "__main__":
    excel_file = r'C:\Users\Nikhil Sharma\Desktop\Internship\Tkinter azeeb.xlsx'
    divide_columns_by_design_area(excel_file)
