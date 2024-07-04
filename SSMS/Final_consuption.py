import os
import pandas as pd
import pyodbc
from openpyxl import load_workbook

def table_exists(cursor, table_name):
    """
    Check if a table exists in the database.
    """
    query = f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table_name}'"
    cursor.execute(query)
    return cursor.fetchone()[0] == 1

def read_excel_with_merged_cells(file_path):
    """
    Read an Excel file with merged cells and return a DataFrame.
    """
    def resolve_merged_ranges(ws):
        """
        Resolve merged cell ranges in a worksheet and return a dictionary mapping merged ranges to their values.
        """
        merged_ranges = {}
        for merged_cell_range in ws.merged_cells.ranges:
            merged_range = []
            min_row, min_col, max_row, max_col = merged_cell_range.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_value = ws.cell(row, col).value
                    merged_range.append(cell_value)
            merged_ranges[(min_row, min_col)] = merged_range
        return merged_ranges

    wb = load_workbook(file_path, data_only=True)
    df = pd.DataFrame()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        merged_ranges = resolve_merged_ranges(ws)
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data_row = []
            for cell_value in row:
                if isinstance(cell_value, tuple) and cell_value in merged_ranges:
                    data_row.append(merged_ranges[cell_value][0])  
                else:
                    data_row.append(cell_value)
            data.append(data_row)
        
        df = pd.DataFrame(data, columns=data[0])  
        break  # Only read the first sheet

    return df

def make_columns_unique(df):
    """
    Function to make column names unique by appending suffixes where necessary.
    """
    seen_columns = {}
    new_columns = []

    for col in df.columns:
        if col in seen_columns:
            suffix = 1
            new_col = f"{col}_{suffix}"
            while new_col in seen_columns:
                suffix += 1
                new_col = f"{col}_{suffix}"
            seen_columns[new_col] = 0
        else:
            new_col = col
            seen_columns[new_col] = 0

        new_columns.append(new_col)

    return new_columns

def fill_empty_headers(headers):
    """
    Fill empty headers with the previous non-empty header.
    """
    filled_headers = []
    previous_header = ""

    for header in headers:
        if header:
            previous_header = header
            filled_headers.append(header)
        else:
            filled_headers.append(previous_header)

    return filled_headers

def concatenate_headers(headers):
    """
    Concatenate headers from row 2 into row 1, replacing existing suffixes.
    """
    concatenated_headers = []
    for header1, header2 in zip(headers[0], headers[1]):
        if header1 and header2:
            concatenated_headers.append(f"{header1} {header2}")
        else:
            concatenated_headers.append(header1 or header2)

    return concatenated_headers

def import_excel_to_sql(input_folder, server_name, database_name):
    """
    Import Excel files from a folder into SQL Server database tables.
    """
    connection_string = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server_name};"
        f"DATABASE={database_name};"
        f"Trusted_Connection=yes;"
    )

    try:
        connection = pyodbc.connect(connection_string)
        cursor = connection.cursor()

        for file_name in os.listdir(input_folder):
            if file_name.startswith('~$'):
                continue 

            file_path = os.path.join(input_folder, file_name)
            
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                df = read_excel_with_merged_cells(file_path)
            else:
                continue  

            table_name = os.path.splitext(file_name)[0]
            table_name = table_name.replace(' ', '_').replace('-', '_').replace('.', '_')

            if table_exists(cursor, table_name):
                continue 

            # Stage 1: Fill empty headers with previous non-empty headers
            filled_headers = fill_empty_headers(df.iloc[0].fillna('').tolist())

            # Stage 2: Concatenate headers from row 2 into row 1
            headers = concatenate_headers([filled_headers, df.iloc[1].fillna('').tolist()])

            df.columns = headers

            # Stage 3: Ensure all column names are unique with numeric suffixes
            new_columns = make_columns_unique(df)
            df.columns = new_columns

            # Stage 4: Create SQL table and insert data
            columns_definition = ", ".join(f"[{col}] VARCHAR(MAX)" for col in df.columns)
            create_table_query = f"CREATE TABLE {table_name} ({columns_definition})"
            cursor.execute(create_table_query)

            for index, row in df.iloc[2:].iterrows():  
                row = ["" if pd.isna(x) else x for x in row]
                insert_query = f"INSERT INTO {table_name} ({', '.join(f'[{col}]' for col in df.columns)}) VALUES ({', '.join(['?']*len(df.columns))})"
                cursor.execute(insert_query, tuple(row))

            connection.commit()

        cursor.close()
        connection.close()
        print("Files imported successfully to SQL Server.")

    except Exception as e:
        print(f"Failed to import files to SQL Server: {e}")

if __name__ == "__main__":
    input_folder = r"C:\Users\Nikhil Sharma\Desktop\Internship"  # Replace with your input folder path
    server_name = 'DESKTOP-PLHJ2M7'
    database_name = 'test'
    
    import_excel_to_sql(input_folder, server_name, database_name)
