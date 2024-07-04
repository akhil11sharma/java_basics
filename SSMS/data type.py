import openpyxl #lib to acess excel
import csv
import pyodbc#for excessing SSMS Server
import os

def parse_excel_or_csv(file_path):
    tables = []
    file_name = os.path.basename(file_path)
    table_name = os.path.splitext(file_name)[0]  # Extract file name without extension
    
    if file_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        columns = [cell.value for cell in ws[1]]  # Assuming headers are in the first row
        tables.append((table_name, columns))
    
    elif file_path.endswith('.csv'):
        with open(file_path, 'r') as csvfile:
            reader = csv.reader(csvfile)
            columns = next(reader)  
            tables.append((table_name, columns))
    
    return tables

def create_table_structure(tables, connection_string, file_path):
    conn = None

    try:
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()

        cursor.execute('''
            IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'datatype')
            BEGIN
                CREATE TABLE datatype (
                    TableName NVARCHAR(MAX),
                    ColumnName NVARCHAR(MAX),
                    DataType NVARCHAR(MAX)
                );
            END
        ''')

        for table_name, columns in tables:
            for column_name in columns:
                sample_value = None
                if file_path.endswith('.xlsx'):#loading of excel file
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    ws = wb.active
                    sample_value = ws.cell(row=2, column=columns.index(column_name) + 1).value
                elif file_path.endswith('.csv'):#loading of CSV File
                    with open(file_path, 'r') as csvfile:
                        reader = csv.reader(csvfile)
                        next(reader)  # Skip header
                        sample_value = next(reader)[columns.index(column_name)]

                data_type = get_data_type(sample_value)
                cursor.execute("INSERT INTO datatype (TableName, ColumnName, DataType) VALUES (?, ?, ?)",
                               table_name, column_name, data_type)

        conn.commit()
        print("Data inserted into 'datatype' table successfully.")#output after submission
    
    except Exception as e:
        print(f"Error: {str(e)}")
    
    finally:
        if conn:
            conn.close()

def get_data_type(sample_value):#datatype description for accurate answers
    try:
        if sample_value is None:
            return 'NVARCHAR(MAX)'

        # Check for boolean
        if isinstance(sample_value, str):
            sample_value_lower = sample_value.lower()
            if sample_value_lower in ['true', 'false']:
                return 'BIT'
        
        if isinstance(sample_value, str) and sample_value.isdigit():
            sample_value = int(sample_value)
        
        if isinstance(sample_value, bool):
            return 'BIT'
        elif isinstance(sample_value, int):
            if -32768 <= sample_value <= 32767:
                return 'SMALLINT'
            elif -2147483648 <= sample_value <= 2147483647:
                return 'INT'
            elif 0 <= sample_value <= 255:
                return 'TINYINT'
            else:
                return 'BIGINT'
        elif isinstance(sample_value, float):
            return 'FLOAT'
        else:
            return 'NVARCHAR(MAX)'
    except Exception as e:
        print(f"Error in determining data type: {str(e)}")
        return 'NVARCHAR(MAX)'

def main():
    file_path = r"C:\Users\Nikhil Sharma\Desktop\SSMS\Calculation.xlsx"#Replace with your file path
    connection_string = 'DRIVER={SQL Server};SERVER=DESKTOP-PLHJ2M7;DATABASE=test;Trusted_Connection=yes;'#Server information 

    tables = parse_excel_or_csv(file_path)
    create_table_structure(tables, connection_string, file_path)

if __name__ == "__main__":
    main()