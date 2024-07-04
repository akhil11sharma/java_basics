import openpyxl

def print_excel_column_headers(file_path):
    workbook = openpyxl.load_workbook(file_path)
    
    print(f"Sheet Names: {workbook.sheetnames}")
    
    for sheet_name in workbook.sheetnames:
        print(f"Sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        headers = next(sheet.iter_rows(values_only=True))
        for header in headers:
            print(header, end='\t')
        print()  

if __name__ == "__main__":
    excel_file = r'C:\Users\Nikhil Sharma\Desktop\Internship\Tkinter azeeb.xlsx'
    print_excel_column_headers(excel_file)
