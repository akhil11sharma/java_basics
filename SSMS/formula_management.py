import openpyxl
import pyodbc
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def extract_formulas(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    formulas = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows())]

        for row_index, row in enumerate(sheet.iter_rows(min_row=2)):
            for col_index, cell in enumerate(row):
                if cell.data_type == 'f':
                    formula_info = {
                        'Formula': cell.value,
                        'Header': headers[col_index],
                        'CellName': cell.coordinate,
                        'InvolvedColumns': replace_with_headers(cell.value, headers, workbook, sheet_name)
                    }
                    formulas.append(formula_info)

    workbook.close()

    return formulas

def replace_with_headers(formula, headers, workbook, current_sheet_name):
    column_references = re.findall(r'([A-Za-z]+)(\d+)', formula)
    replaced_formula = formula

    for col_ref, row_num in column_references:
        if col_ref.isalpha():  # Check if it's a valid column reference
            column_name = col_ref.upper()
            try:
                col_index = openpyxl.utils.cell.column_index_from_string(column_name) - 1

                if col_index < len(headers) and headers[col_index]:
                    header_name = headers[col_index]
                    replaced_formula = replaced_formula.replace(f"{col_ref}{row_num}", header_name)
                else:
                    # Handle cross-sheet references
                    replaced_formula = replaced_formula.replace(f"{col_ref}{row_num}", f"[{workbook[current_sheet_name].title}]!")
            except ValueError:
                # Handle cases where column_index_from_string fails (invalid column reference)
                # This may happen if the reference is not a valid column but something else
                # Try to find the reference in other sheets
                for sheet in workbook.sheetnames:
                    if f"{sheet}!" in formula:
                        replaced_formula = replaced_formula.replace(f"{sheet}!", f"[{workbook[sheet].title}]!")
        else:
            # Handle cases where the reference is not a column but possibly a sheet name or invalid reference
            # Try to find the reference in other sheets
            for sheet in workbook.sheetnames:
                if f"{sheet}!" in formula:
                    replaced_formula = replaced_formula.replace(f"{sheet}!", f"[{workbook[sheet].title}]!")

    return replaced_formula

def table_exists(cursor, table_name):
    cursor.execute(f'''
        SELECT COUNT(*)
        FROM information_schema.tables
        WHERE table_name = '{table_name}'
    ''')
    return cursor.fetchone()[0] == 1

def create_tracking_table_if_not_exists(cursor):
    cursor.execute(f'''
        IF NOT EXISTS (SELECT * FROM sys.tables WHERE name='areaCalculationSheet_Formulas')
        BEGIN
            CREATE TABLE areaCalculationSheet_Formulas (
                Formula NVARCHAR(MAX),
                Column_Header NVARCHAR(255),
                CellName NVARCHAR(255),
                Involved_Columns NVARCHAR(MAX),
                Table_Name NVARCHAR(255)
            );
            PRINT 'Tracking table created successfully: areaCalculationSheet_Formulas';
        END
    ''')

def update_existing_table(cursor, table_name, formulas):
    for formula_info in formulas:
        formula = formula_info['Formula']
        header = formula_info['Header']
        cell_name = formula_info['CellName']
        involved_columns = formula_info['InvolvedColumns']

        cursor.execute(f'''
            IF EXISTS (SELECT * FROM areaCalculationSheet_Formulas WHERE CellName = ? AND Table_Name = ?)
            BEGIN
                UPDATE areaCalculationSheet_Formulas
                SET Formula = ?, Column_Header = ?, Involved_Columns = ?
                WHERE CellName = ? AND Table_Name = ?
            END
            ELSE
            BEGIN
                INSERT INTO areaCalculationSheet_Formulas (Formula, Column_Header, CellName, Involved_Columns, Table_Name)
                VALUES (?, ?, ?, ?, ?)
            END
        ''', (cell_name, table_name, formula, header, involved_columns, cell_name, table_name, formula, header, cell_name, involved_columns, table_name))

def create_view_if_not_exists(cursor):
    cursor.execute('''
        IF OBJECT_ID('areaCalculationSheet_Formulas_View', 'V') IS NULL
        BEGIN
            EXEC sp_executesql N'
                CREATE VIEW areaCalculationSheet_Formulas_View AS
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS [S.NO],
                    CellName AS [Cell_Name],
                    Column_Header AS [Column_Header],
                    Formula AS [Formula],
                    Involved_Columns AS [Involved_Formula],
                    Table_Name AS [Table_Name]
                FROM 
                    areaCalculationSheet_Formulas
            ';
            PRINT 'View created successfully: areaCalculationSheet_Formulas_View';
        END
    ''')

def store_in_database(formulas, table_name, db_config):
    conn_str = f"DRIVER={{SQL Server}};SERVER={db_config['server']};DATABASE={db_config['database']};trusted_connection={db_config['trusted_connection']}"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    create_tracking_table_if_not_exists(cursor)
    update_existing_table(cursor, table_name, formulas)
    create_view_if_not_exists(cursor)

    cursor.commit()
    conn.close()

def process_file(file_path, db_config):
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    formulas = extract_formulas(file_path)

    if formulas:
        store_in_database(formulas, file_name, db_config)
        return f"File '{file_path}' successfully processed and data stored in table 'areaCalculationSheet_Formulas' under '{file_name}'."
    else:
        return "No formulas found in the Excel sheet."

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        selected_file_label.config(text=f"Selected File: {file_path}")
        execute_button.pack()

def execute_process():
    file_path = selected_file_label.cget("text").split(": ")[1]  # Get the selected file path
    message = process_file(file_path, db_config)
    messagebox.showinfo("Process Completed", message)

def display_initial_screen():
    for widget in frame.winfo_children():
        widget.destroy()

    select_label = ttk.Label(frame, text="SELECT THE FILE TO SEND", font=("Helvetica", 16, "bold"))
    select_label.pack(pady=10)

    select_button = ttk.Button(frame, text="Select Excel File", style="Accent.TButton", command=select_file)
    select_button.pack(pady=10)

    global selected_file_label
    selected_file_label = ttk.Label(frame, text="Selected File: None")
    selected_file_label.pack(pady=10)

    global execute_button
    execute_button = ttk.Button(frame, text="Execute", style="Accent.TButton", command=execute_process)

if __name__ == "__main__":
    db_config = {
        'server': 'DESKTOP-PLHJ2M7',
        'database': 'test',
        'trusted_connection': 'yes'
    }

    root = tk.Tk()
    root.title("Formula Extraction")
    root.geometry("600x300")

    frame = ttk.Frame(root, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    style = ttk.Style()
    style.configure("Accent.TButton", foreground="black", background="orange", font=("Helvetica", 12))

    display_initial_screen()

    root.mainloop()
